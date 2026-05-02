from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl import Workbook
import os
import time
from datetime import date, datetime
import calendar
import threading

EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'data', 'timesheet.xlsx')
app = Flask(__name__)
lock = threading.Lock()


def str_date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime('%Y-%m-%d')
    s = str(val)
    return s[:10] if len(s) >= 10 else None


def init_workbook():
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'metadata'
    ws.append(['month', 'category', 'daily_target'])
    wb.create_sheet('entries').append(['date', 'category', 'hours'])
    wb.create_sheet('leaves').append(['date'])
    wb.save(EXCEL_FILE)
    return wb


def get_workbook():
    if not os.path.exists(EXCEL_FILE):
        return init_workbook()
    try:
        return openpyxl.load_workbook(EXCEL_FILE)
    except Exception:
        os.remove(EXCEL_FILE)
        return init_workbook()


def save_workbook(wb):
    # Retry up to 10 times — OneDrive briefly locks the file during sync
    for attempt in range(10):
        try:
            wb.save(EXCEL_FILE)
            return
        except PermissionError:
            if attempt < 9:
                time.sleep(0.3)
            else:
                raise PermissionError(
                    "Cannot save — close timesheet.xlsx in Excel if it is open"
                )


# ─── METADATA ────────────────────────────────────────────────────────────────

@app.route('/api/metadata', methods=['GET'])
def get_metadata():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    with lock:
        wb = get_workbook()
        rows = list(wb['metadata'].iter_rows(min_row=2, values_only=True))

    categories = [
        {'category': r[1], 'daily_target': float(r[2] or 0)}
        for r in rows if r[0] == month
    ]

    # If this month has no categories yet, inherit from the most recent past month
    inherited = False
    if not categories:
        months_with_data = sorted({r[0] for r in rows if r[0] and r[0] < month}, reverse=True)
        if months_with_data:
            source = months_with_data[0]
            categories = [
                {'category': r[1], 'daily_target': float(r[2] or 0)}
                for r in rows if r[0] == source
            ]
            inherited = True

    return jsonify({'month': month, 'categories': categories, 'inherited': inherited})


@app.route('/api/metadata', methods=['POST'])
def save_metadata():
    data = request.json
    month = data['month']
    categories = data['categories']
    with lock:
        wb = get_workbook()
        ws = wb['metadata']
        to_del = [i for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2)
                  if r[0] == month]
        for i in reversed(to_del):
            ws.delete_rows(i)
        for c in categories:
            ws.append([month, c['category'], float(c['daily_target'])])
        save_workbook(wb)
    return jsonify({'success': True})


# ─── ENTRIES ─────────────────────────────────────────────────────────────────

@app.route('/api/entries', methods=['GET'])
def get_entries():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    with lock:
        wb = get_workbook()
        rows = list(wb['entries'].iter_rows(min_row=2, values_only=True))
    entries = {}
    for r in rows:
        d = str_date(r[0])
        if d and d[:7] == month:
            entries.setdefault(d, {})[r[1]] = float(r[2] or 0)
    return jsonify({'month': month, 'entries': entries})


@app.route('/api/entries', methods=['POST'])
def save_entry():
    data = request.json
    date_str, category, hours = data['date'], data['category'], float(data['hours'])
    with lock:
        wb = get_workbook()
        ws = wb['entries']
        updated = False
        for row in ws.iter_rows(min_row=2):
            if str_date(row[0].value) == date_str and row[1].value == category:
                row[2].value = hours
                updated = True
                break
        if not updated:
            ws.append([date_str, category, hours])
        save_workbook(wb)
    return jsonify({'success': True})


@app.route('/api/entries', methods=['DELETE'])
def delete_entry():
    date_str = request.args.get('date')
    category = request.args.get('category')
    with lock:
        wb = get_workbook()
        ws = wb['entries']
        to_del = [i for i, row in enumerate(ws.iter_rows(min_row=2), 2)
                  if str_date(row[0].value) == date_str and row[1].value == category]
        for i in reversed(to_del):
            ws.delete_rows(i)
        save_workbook(wb)
    return jsonify({'success': True})


# ─── LEAVES ──────────────────────────────────────────────────────────────────

@app.route('/api/leaves', methods=['GET'])
def get_leaves():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    with lock:
        wb = get_workbook()
        rows = list(wb['leaves'].iter_rows(min_row=2, values_only=True))
    leaves = [str_date(r[0]) for r in rows if r[0] and str_date(r[0]) and str_date(r[0])[:7] == month]
    return jsonify({'month': month, 'leaves': leaves})


@app.route('/api/leaves', methods=['POST'])
def save_leaves():
    data = request.json
    month, leaves = data['month'], data['leaves']
    with lock:
        wb = get_workbook()
        ws = wb['leaves']
        to_del = [i for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2)
                  if r[0] and str_date(r[0]) and str_date(r[0])[:7] == month]
        for i in reversed(to_del):
            ws.delete_rows(i)
        for leaf in leaves:
            ws.append([leaf])
        save_workbook(wb)
    return jsonify({'success': True})


# ─── DATES (for dropdowns) ────────────────────────────────────────────────────

@app.route('/api/dates', methods=['GET'])
def get_dates():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    year_int, mon_int = map(int, month.split('-'))
    with lock:
        wb = get_workbook()
        rows = list(wb['leaves'].iter_rows(min_row=2, values_only=True))
    leaves = {str_date(r[0]) for r in rows if r[0] and str_date(r[0]) and str_date(r[0])[:7] == month}

    total_days = calendar.monthrange(year_int, mon_int)[1]
    result = []
    for day in range(1, total_days + 1):
        d = date(year_int, mon_int, day)
        ds = d.strftime('%Y-%m-%d')
        is_weekend = d.weekday() >= 5
        is_leave = ds in leaves
        result.append({
            'date': ds,
            'label': d.strftime('%d %b %Y (%a)'),
            'is_weekend': is_weekend,
            'is_leave': is_leave,
            'is_working': not is_weekend and not is_leave
        })
    return jsonify({'month': month, 'dates': result})


# ─── SUMMARY ─────────────────────────────────────────────────────────────────

@app.route('/api/summary', methods=['GET'])
def get_summary():
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    today = date.today()
    year_int, mon_int = map(int, month.split('-'))

    with lock:
        wb = get_workbook()
        meta_rows = list(wb['metadata'].iter_rows(min_row=2, values_only=True))
        leave_rows = list(wb['leaves'].iter_rows(min_row=2, values_only=True))
        entry_rows = list(wb['entries'].iter_rows(min_row=2, values_only=True))

    categories = {r[1]: {'daily_target': float(r[2] or 0), 'completed': 0.0}
                  for r in meta_rows if r[0] == month}
    leaves = {str_date(r[0]) for r in leave_rows
              if r[0] and str_date(r[0]) and str_date(r[0])[:7] == month}

    if year_int == today.year and mon_int == today.month:
        end_day = today.day
    else:
        end_day = calendar.monthrange(year_int, mon_int)[1]

    working_days = sum(
        1 for day in range(1, end_day + 1)
        if date(year_int, mon_int, day).weekday() < 5
        and date(year_int, mon_int, day).strftime('%Y-%m-%d') not in leaves
    )

    for r in entry_rows:
        d = str_date(r[0])
        if d and d[:7] == month and r[1] in categories:
            categories[r[1]]['completed'] += float(r[2] or 0)

    for cat, data in categories.items():
        target_so_far = round(data['daily_target'] * working_days, 2)
        completed = round(data['completed'], 2)
        data.update({
            'target_so_far': target_so_far,
            'completed': completed,
            'pending': round(target_so_far - completed, 2)
        })

    return jsonify({'month': month, 'working_days': working_days, 'categories': categories})


# ─── YEARLY ──────────────────────────────────────────────────────────────────

@app.route('/api/yearly', methods=['GET'])
def get_yearly():
    year = int(request.args.get('year', date.today().year))
    today = date.today()

    with lock:
        wb = get_workbook()
        meta_rows = list(wb['metadata'].iter_rows(min_row=2, values_only=True))
        leave_rows = list(wb['leaves'].iter_rows(min_row=2, values_only=True))
        entry_rows = list(wb['entries'].iter_rows(min_row=2, values_only=True))

    result = {}
    for mon_int in range(1, 13):
        if year > today.year or (year == today.year and mon_int > today.month):
            continue
        month = f"{year}-{str(mon_int).zfill(2)}"

        categories = {r[1]: {'daily_target': float(r[2] or 0), 'completed': 0.0}
                      for r in meta_rows if r[0] == month}
        if not categories:
            continue

        leaves = {str_date(r[0]) for r in leave_rows
                  if r[0] and str_date(r[0]) and str_date(r[0])[:7] == month}

        if year == today.year and mon_int == today.month:
            end_day = today.day
        else:
            end_day = calendar.monthrange(year, mon_int)[1]

        working_days = sum(
            1 for day in range(1, end_day + 1)
            if date(year, mon_int, day).weekday() < 5
            and date(year, mon_int, day).strftime('%Y-%m-%d') not in leaves
        )

        for r in entry_rows:
            d = str_date(r[0])
            if d and d[:7] == month and r[1] in categories:
                categories[r[1]]['completed'] += float(r[2] or 0)

        for data in categories.values():
            target = round(data['daily_target'] * working_days, 2)
            data.update({
                'target': target,
                'completed': round(data['completed'], 2),
                'pending': round(target - data['completed'], 2)
            })

        result[month] = {'working_days': working_days, 'categories': categories}

    return jsonify({'year': year, 'months': result})


# ─── METADATA ALL MONTHS ──────────────────────────────────────────────────────

@app.route('/api/metadata/all', methods=['GET'])
def get_metadata_all():
    with lock:
        wb = get_workbook()
        rows = list(wb['metadata'].iter_rows(min_row=2, values_only=True))
    by_month = {}
    for r in rows:
        if r[0] and r[1]:
            by_month.setdefault(r[0], {})[r[1]] = float(r[2] or 0)
    all_cats = sorted({cat for cats in by_month.values() for cat in cats})
    months = sorted(by_month.keys())
    return jsonify({'months': months, 'categories': all_cats, 'data': by_month})


@app.route('/')
def index():
    return render_template('index.html')


if __name__ == '__main__':
    app.run(debug=True, port=5000)
